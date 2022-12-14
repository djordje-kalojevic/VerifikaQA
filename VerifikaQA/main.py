"""Allows user to create Verifika quality assurance reports
without the need for Verifika's UI or various dialogues.
In turn, this speeds up the process."""

from os import mkdir, remove, system
from os.path import isfile, isdir, split
from shutil import rmtree, copy
from tkinter import PhotoImage, TclError, Tk, Toplevel, Button, Checkbutton, StringVar, BooleanVar, Label
from tkinter.messagebox import showinfo, showerror, askyesno
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
from tkinter.ttk import Radiobutton
from configparser import ConfigParser
from subprocess import Popen, DEVNULL, STDOUT
import sys
from psutil import process_iter, NoSuchProcess
from openpyxl import load_workbook
from PIL import Image
from alive_progress import alive_bar


def create_program_mainloop(transparent_icon_location='images/icon.ico') -> Tk:
    """Creates an instance of tk.Tk class and replaces its icon with a transparent one."""

    if not isfile(transparent_icon_location):
        transparent_icon = Image.new("RGBA", (16, 16), (0, 0, 0, 0))
        transparent_icon.save(transparent_icon_location, "ICO")

    root = Tk()
    root.title("")
    # hides its window
    root.withdraw()
    root.iconbitmap(True, transparent_icon_location)

    return root


class ConfigFile(ConfigParser):
    """Custom ConfigParser class that stores its own path.
    This allows it to be read without needing to pass its path to the function.
    Use: ConfigFile.read(ConfigFile.path)"""

    def __init__(self, path):
        ConfigParser.__init__(self)
        self.path = path
        self.read(self.path)


def create_config(config_location='config.ini') -> ConfigFile:
    """Creates a simple config file which will keep the following:
    verifika__location: location of the Verifika executable,
    verifika_profiles_location: location of a directory containing Verifika profiles"""

    config_file = ConfigFile(config_location)
    config_file.read(config_location)

    if not isfile(config_location):
        with open(config_location, "w", encoding="utf-8") as file:
            config_file.write(file)

    return config_file


def update_config(config_file: ConfigFile, section: str, option: str,
                  value: str):
    """General update function which can be used for future features."""

    config_file.set(section, option, value)
    with open(config_file.path, "w", encoding="utf-8") as file:
        config_file.write(file)


class ToggleButton(Button):
    """"Button class that changes its image based on whether it is on or off.
    Additionally it can return its current state (on or off)."""

    def __init__(self, master):
        self.on_image = PhotoImage(file="images/enabled.png")
        self.off_image = PhotoImage(file="images/disabled.png")
        self.manual_optimization = False

        Button.__init__(self, master, image=self.off_image, bd=0)
        self.master = master
        self['command'] = self.toggle

    def toggle(self):
        """Allows the button to be toggled on and off."""

        if not self.manual_optimization:
            self.config(image=self.on_image)
            self.manual_optimization = True

        else:
            self.config(image=self.off_image)
            self.manual_optimization = False

    def check_state(self) -> bool:
        """Returns the state of the toggle button, i.e., enabled or disabled."""

        return self.manual_optimization


def select_report_type(root: Tk) -> tuple[list[str], bool]:
    """Displays radiobuttons which allow the user to choose which type of report they need.
    Furthermore allows the creation of custom reports via checkboxes.
    Returns which types of errors user would like to keep as well as
    wether they want to optimize the report manually."""

    radiobutton_popup = Toplevel(root)
    radiobutton_popup.attributes("-topmost", "true")
    radiobutton_popup.geometry("225x275")

    label = Label(radiobutton_popup,
                  text="\nPlease choose desired report type:\n")
    label.grid(columnspan=3, column=0, row=0, ipadx=20)

    toggle_button_label = Label(radiobutton_popup, text='Manual Optimization')
    toggle_button_label.grid(columnspan=2, column=0, row=1, ipady=10, ipadx=20)

    toggle_button = ToggleButton(radiobutton_popup)
    toggle_button.grid(column=2, row=1, ipady=0)

    str_var = StringVar(radiobutton_popup, "1")

    # dictionary containing labels and return values for radiobuttons
    radiobutton_dictionary = {
        "Full Report": "Full",
        "Consistency Report": "Consistency Errors",
        "Spelling + Grammar Report": "Spelling + Grammar",
        "Custom Report": "Custom"
    }

    # iterating over labels and adding buttons to the window
    row_number = 2
    for (text, value) in radiobutton_dictionary.items():
        radiobutton = Radiobutton(radiobutton_popup,
                                  text=text,
                                  variable=str_var,
                                  value=value)
        radiobutton.grid(columnspan=3, column=0, row=row_number, ipady=10)
        row_number += 1

    # waits for the user input
    radiobutton_popup.wait_variable(str_var)

    # get the label of a radiobutton selected
    choice = str(str_var.get())
    sheets_to_keep = []

    if choice == "Custom":
        sheets_to_keep = checkbuttons_window(root)

    #custom spellcheck QA check
    #custom spellcheck QA check
    elif choice == "Spelling + Grammar":
        sheets_to_keep.extend(
            ["Spelling Errors", "Grammar Errors", "User-defined Errors"])

    else:
        sheets_to_keep.append(choice)

    manual_optimization = toggle_button.check_state()

    #closes the window if it remained open
    try:
        radiobutton_popup.destroy()
    except TclError:
        pass

    return sheets_to_keep, manual_optimization


class CheckBox(Checkbutton):
    """Custom tk.Checkbutton class that stores labels of checked buttons"""

    # Storage for all buttons
    boxes: list[Checkbutton] = []

    def __init__(self, master=None, **options):
        Checkbutton.__init__(self, master, options)
        self.boxes.append(self)
        # var used to store checkbox state (on/off)
        self.var = BooleanVar()
        # allows the storage of its label
        self.text = self.cget("text")
        # set the checkbox to use the new var
        self.configure(variable=self.var)


def checkbuttons_window(root: Tk) -> list[str]:
    """Supports custom reports via checkbuttons"""

    # list containing labels for checkbuttons
    checkbutton_list = [
        "Common Errors", "Consistency Errors", "Spelling Errors",
        "Grammar Errors", "User-Defined Errors"
    ]

    checkbox_popup = Toplevel(root)
    checkbox_popup.attributes("-topmost", "true")
    checkbox_popup.geometry("250x265")
    label = Label(checkbox_popup,
                  text="\nPlease select types of errors to include:\n")
    label.pack()

    # iterates over values and adds buttons to the window
    for button_label in checkbutton_list:
        button = CheckBox(checkbox_popup, text=button_label)
        button.pack(pady=3)

    def check_all():
        for box in CheckBox.boxes:
            box.select()

    def uncheck_all():
        for box in CheckBox.boxes:
            box.var.set(False)

    sheets_to_keep = []

    def confirm():
        for box in CheckBox.boxes:
            if box.var.get():  # Checks if the button is ticked
                sheets_to_keep.append(box.text)

        # makes button non-interactive until at least one option is selected
        if len(sheets_to_keep) > 0:
            root.destroy()

    check_all_button = Button(checkbox_popup,
                              text="Check all",
                              command=check_all,
                              width=10)
    check_all_button.pack(side="left")

    uncheck_all_button = Button(checkbox_popup,
                                text="Uncheck all",
                                command=uncheck_all,
                                width=10)
    uncheck_all_button.pack(side="right")

    confirm_button = Button(checkbox_popup,
                            text="Confirm",
                            command=confirm,
                            width=10)
    confirm_button.pack(pady=16)

    # 'X' closes this program
    checkbox_popup.protocol("WM_DELETE_WINDOW", sys.exit)
    checkbox_popup.mainloop()

    return sheets_to_keep


def browse_verifika(root: Tk, config_file: ConfigFile) -> str:
    """Browses for the executable and writes its location to the config file."""

    close_verifika()

    # tries to read location from config, otherwise let user browse for the executable
    # then updates config file to reflect this
    try:
        verifika_exe_location = config_file["DEFAULT"]["verifika_location"]
        if not isfile(verifika_exe_location):
            raise FileNotFoundError

    except (KeyError, FileNotFoundError):
        verifika_exe_location = askopenfilename(
            parent=root,
            title="Navigate to Verifika directory",
            filetypes=[("Verifika executable", "verifika.exe")])

        if not verifika_exe_location:
            sys.exit()

        update_config(config_file, "DEFAULT", "verifika_location",
                      verifika_exe_location)

    return verifika_exe_location


def browse_verifika_profile(root: Tk, config_file: ConfigFile) -> str:
    """Browses for a Verifika profile and writes its directory to the config file."""

    # tries to read location from config
    try:
        profiles_directory = config_file["DEFAULT"][
            "verifika_profiles_location"]
    except KeyError:
        profiles_directory = ""

    # prompts user to select profile with initial folder location if available
    verifika_profile = askopenfilename(parent=root,
                                       initialdir=profiles_directory,
                                       title="Choose a profile",
                                       filetypes=[
                                           ("Verifika profile (.vprofile)",
                                            "*.vprofile")
                                       ])

    if not verifika_profile:
        sys.exit()

    # writes location to config if it does not exist
    if profiles_directory == "" or not isdir(profiles_directory):
        profiles_directory = split(verifika_profile)[0]

        update_config(config_file, "DEFAULT", "verifika_profiles_location",
                      profiles_directory)

    return verifika_profile


def close_verifika():
    """Checks if Verifika is already running and prompts user to close it."""

    for process in process_iter(attrs=["name"]):
        if process.name() == "Verifika.exe":

            answer = askyesno(
                title="Verifika is already running",
                message=("Another instance of Verifika is already running.\n"
                         "Would you like to close it before continuing?\n"
                         "Warning: no changes will be saved."),
                icon="question")

            # closes this program if user selects "no" (or closes the window)
            if not answer:
                sys.exit()

            try:
                process.kill()

            # if user closes the program before selecting "yes"
            except NoSuchProcess:
                pass


def select_files(root: Tk) -> tuple[tuple[str, ...], str]:
    """Prompts user to select files for checking.
    Returns the list of files and their directory."""

    files = askopenfilenames(parent=root,
                             title="Choose a file",
                             filetypes=[("Any .xliff file", "*.*xliff")])

    files_dir = split(files[0])[0]

    # program shuts down if no files have been selected
    if not files:
        sys.exit()

    return files, files_dir


def manage_files(files: tuple[str]) -> str:
    """Returns file(s) location and their path.\n
    Verifika's CMD implementation has an inbuilt limitation and
    cannot load in multiple files at once,
    this was circumvented by copying them to a subdirectory and loading it instead.
    Therefore this function returns either the file's or directory's path.\n
    Note: This is a separate function in order to make UI more responsive,
    i.e., user can select all files and options before any intensive work is performed."""

    # gets the directory where files are located
    files_dir = split(files[0])[0]
    temp_dir = f"{files_dir}/temp_dir"

    if len(files) > 1:
        # removes sub-dir if it already exists
        if isdir(temp_dir):
            rmtree(temp_dir)

        # creates sub-dir and copies files to it
        mkdir(temp_dir)
        source_dir = files_dir + "\\"
        dst_sir = temp_dir + "\\"

        for file in files:
            # gets the name of the file
            file_name = split(file)[1]
            file_source_location = source_dir + file_name
            file_dst_location = dst_sir + file_name
            copy(file_source_location, file_dst_location)

        files_to_check = temp_dir

    # if there is only one file to be checked
    else:
        files_to_check = files[0]

    return files_to_check


def run_qa(files_dir: str, verifika_exe_location: str, files_to_check: str,
           verifika_profile: str, sheets_to_keep: list[str],
           manual_optimization: bool):
    """Preforms Verifika QA via CMD.\n
    Returns QA report's name, if one was saved."""

    temp_report_name = f"{files_dir}/temp_report.xlsx"

    # more optimized reports for individual Common, Consistency, and Spelling errors.
    condition = ["Common Errors", "Consistency Errors", "Spelling Errors"]
    if len(sheets_to_keep) == 1 and any(sheet in sheets_to_keep
                                        for sheet in condition):
        # takes just the first word
        choice = sheets_to_keep[0]
        report_type = choice.split(" ")[0]
        cmd_command = (
            f'"{verifika_exe_location}" -files "{files_to_check}" -profile "{verifika_profile}"'
            f' -startcheck -type {report_type}')

    # otherwise uses general Full report mode
    else:
        cmd_command = (
            f'"{verifika_exe_location}" -files "{files_to_check}" -profile "{verifika_profile}"'
            f' -startcheck -type Full')

    # automatic optimization additionally requires path in order to store the report once created
    if not manual_optimization:
        cmd_command += f' -result {temp_report_name}'

    # runs QA via CMD with accompanying progress bar
    # DEVNULL used to suppress the long output from Verifika, error output left enabled
    with alive_bar(spinner=None,
                   title="Performing QA check",
                   stats=False,
                   monitor=None,
                   monitor_end='Performed in') as progress_bar:
        with Popen(cmd_command, stdout=DEVNULL, stderr=STDOUT) as proc:
            # waits for command line to finish QA
            proc.wait()
            # stops progress bar and shows elapsed time
            progress_bar()  # pylint: disable=not-callable

    # deletes temp-dir as it is no longer necessary
    temp_dir = f"{files_dir}/temp_dir"
    if isdir(temp_dir):
        rmtree(temp_dir)

    if not isfile(temp_report_name):
        if not manual_optimization:
            showinfo(title="No report saved", message="No errors were found.")
            sys.exit()
    else:
        process_and_save_report(temp_report_name, sheets_to_keep)


def process_and_save_report(temp_report_name: str, sheets_to_keep: list[str]):
    """Note: Sometimes Excel reports issues when opening these files,, however,
    they can be fixed."""

    files_dir = split(temp_report_name)[0]
    wb_verifika = load_workbook(temp_report_name, read_only=False)
    remove(temp_report_name)

    for sheet_name in wb_verifika.sheetnames:
        sheet = wb_verifika[sheet_name]
        # removes sheets that are just Verifika sheet headers
        if sheet.max_row < 12:
            wb_verifika.remove(sheet)
        # further optimization for non-Full reports
        elif "Full" not in sheets_to_keep and sheet_name not in sheets_to_keep:
            wb_verifika.remove(sheet)

    if len(wb_verifika.sheetnames) > 0:
        report_saved = False
        while not report_saved:
            try:
                new_report_name = asksaveasfilename(initialdir=files_dir,
                                                    initialfile="QA report",
                                                    defaultextension=".xlsx",
                                                    filetypes=[
                                                        ("Excel file (.xlsx)",
                                                         "*.xlsx")
                                                    ])

                if not new_report_name:
                    sys.exit()

                report_saved = True
            except PermissionError:
                showerror(
                    title="Error occurred!",
                    message=
                    ("File could not be saved because it is already opened "
                     "by another process (most likely Excel or another reader). "
                     "Please close it before continuing."))

        wb_verifika.save(new_report_name)

        # opens saved report
        system(f'"{new_report_name}"')


def main():
    """Allows user to create Verifika reports without the need for its UI."""

    root = create_program_mainloop()

    config_file = create_config()

    verifika_exe_location = browse_verifika(root, config_file)

    files, files_dir = select_files(root)

    verifika_profile = browse_verifika_profile(root, config_file)

    sheets_to_keep, manual_optimization = select_report_type(root)

    files_to_check = manage_files(files)

    run_qa(files_dir, verifika_exe_location, files_to_check, verifika_profile,
           sheets_to_keep, manual_optimization)


if __name__ == "__main__":
    main()
